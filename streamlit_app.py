import streamlit as st
import pickle
import numpy as np
import pandas as pd
import joblib
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Configuração da página
st.set_page_config(
    page_title="Previsão de Resistência do Concreto",
    page_icon="https://raw.githubusercontent.com/dimasbetioli/concrete-ia-app/refs/heads/main/mult.png",
    layout="centered",
    initial_sidebar_state="expanded"
)

# Adicionar imagem no topo da página
st.image(
    "https://raw.githubusercontent.com/dimasbetioli/concrete-ia-app/refs/heads/main/topo.png",
    use_container_width=True
)

# Título principal
st.markdown(
    "<h1 style='text-align: center; color: #2196F3; font-size: 30px;'> PROJETO - IA APLICADA À PREVISÃO DA RESISTÊNCIA DE CONCRETOS AOS 28 DIAS - VERSÃO BETA</h1>",
    unsafe_allow_html=True,
)

# Introdução
st.markdown(
    '<h3 style="text-align: center; color: #2C2C2C; font-size: 14px; margin-bottom: 20px;">Equipe liderada pelo Prof André C.P.L.F. de Carvalho</h3>',
    unsafe_allow_html=True,
)

# Opções de entrada: manual ou por arquivo
# Título informativo
st.markdown(
    """
    <p style="text-align: center; font-size: 16px; color: #2C2C2C;">
        <span style="font-size: 30px; color: #4CAF50;">&#8595;</span> Escolha se deseja inserir os dados manualmente ou carregar um arquivo Excel <span style="font-size: 30px; color: #4CAF50;">&#8595;</span>
    </p>
    """,
    unsafe_allow_html=True
)

# Inicializando o estado da escolha
if 'tipo_entrada' not in st.session_state:
    st.session_state.tipo_entrada = None

# Centralizar e exibir botões lado a lado
col1, col2 = st.columns(2)  # Criando duas colunas de igual largura para exibir os botões lado a lado

with col1:
    if st.button("Inserir manualmente"):
        st.session_state.tipo_entrada = "Inserir manualmente"

with col2:
    if st.button("Carregar arquivo Excel"):
        st.session_state.tipo_entrada = "Carregar arquivo Excel"

# Adicionar espaço acima do botão
st.markdown("<div style='margin-top: 30px;'></div>", unsafe_allow_html=True)

# Lógica condicional baseada na escolha do usuário
if st.session_state.tipo_entrada == "Inserir manualmente":
    st.write("Você escolheu inserir os dados manualmente.")

    # Opções de configuração
    opcao = st.radio(
        "Escolha a configuração de entrada:",
        [
            "CT_Cimento e CT_Agua",
            "CT_Cimento, CT_Agua, e resistências reais (3d, 7d, 28d)",
            "CT_Cimento, CT_Agua, resistências reais, e Fc_7d",
            "CT_Cimento, CT_Agua, resistências reais, Fc_7d, e aditivos",
            "Todas as variáveis"
        ]
    )

    entradas = []
    model_path = ""

    # --- Session State for Inputs ---
    if 'entradas' not in st.session_state:
        st.session_state.entradas = []

    if opcao == "CT_Cimento e CT_Agua":
        model_path = "modelo1.pkl"
        st.session_state.entradas = [
            st.number_input("CT_Cimento (kg/m³):", min_value=0.0, step=1.0, key="ct_cimento"),
            st.number_input("CT_Agua (kg/m³):", min_value=0.0, step=1.0, key="ct_agua")
        ]
    
    elif opcao == "CT_Cimento, CT_Agua, e resistências reais (3d, 7d, 28d)":
        model_path = "modelo2.pkl"
        st.session_state.entradas = [
            st.number_input("CT_Cimento (kg/m³):", min_value=0.0, step=1.0, key="ct_cimento"),
            st.number_input("CT_Agua (kg/m³):", min_value=0.0, step=1.0, key="ct_agua"),
            st.number_input("cimento_Resistencia_real_3d (MPa):", min_value=0.0, step=1.0, key="resistencia_3d"),
            st.number_input("cimento_Resistencia_real_7d (MPa):", min_value=0.0, step=1.0, key="resistencia_7d"),
            st.number_input("cimento_Resistencia_real_28d (MPa):", min_value=0.0, step=1.0, key="resistencia_28d")
        ]
    
    elif opcao == "CT_Cimento, CT_Agua, resistências reais, e Fc_7d":
        model_path = "modelo3.pkl"
        st.session_state.entradas = [
            st.number_input("CT_Cimento (kg/m³):", min_value=0.0, step=1.0, key="ct_cimento"),
            st.number_input("CT_Agua (kg/m³):", min_value=0.0, step=1.0, key="ct_agua"),
            st.number_input("cimento_Resistencia_real_3d (MPa):", min_value=0.0, step=1.0, key="resistencia_3d"),
            st.number_input("cimento_Resistencia_real_7d (MPa):", min_value=0.0, step=1.0, key="resistencia_7d"),
            st.number_input("cimento_Resistencia_real_28d (MPa):", min_value=0.0, step=1.0, key="resistencia_28d"),
            st.number_input("Fc_7d (MPa):", min_value=0.0, step=1.0, key="fc_7d")
        ]
    
    elif opcao == "CT_Cimento, CT_Agua, resistências reais, Fc_7d, e aditivos":
        model_path = "modelo4.pkl"
        st.session_state.entradas = [
            st.number_input("CT_Cimento (kg/m³):", min_value=0.0, step=1.0, key="ct_cimento"),
            st.number_input("CT_Agua (kg/m³):", min_value=0.0, step=1.0, key="ct_agua"),
            st.number_input("cimento_Resistencia_real_3d (MPa):", min_value=0.0, step=1.0, key="resistencia_3d"),
            st.number_input("cimento_Resistencia_real_7d (MPa):", min_value=0.0, step=1.0, key="resistencia_7d"),
            st.number_input("cimento_Resistencia_real_28d (MPa):", min_value=0.0, step=1.0, key="resistencia_28d"),
            st.number_input("Fc_7d (MPa):", min_value=0.0, step=1.0, key="fc_7d"),
            st.number_input("CT_Silica (kg/m³):", min_value=0.0, step=1.0, key="ct_silica"),
            st.number_input("CT_Plastificante (kg/m³):", min_value=0.0, step=1.0, key="ct_plastificante"),
            st.number_input("CT_Polifuncional (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Superplastificante (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Brita_0 (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Brita_1 (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Areia_natural (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Areia_artificial (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_AC (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Aditivo (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Teor_de_Argamassa (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Teor_de_Agua (kg/m³):", min_value=0.0, step=1.0),
        ]
    elif opcao == "Todas as variáveis":
        model_path = "modelo5.pkl"
        st.session_state.entradas = [
            st.number_input("CT_Cimento (kg/m³):", min_value=0.0, step=1.0, key="ct_cimento"),
            st.number_input("CT_Agua (kg/m³):", min_value=0.0, step=1.0, key="ct_agua"),
            st.number_input("cimento_Resistencia_real_3d (MPa):", min_value=0.0, step=1.0, key="resistencia_3d"),
            st.number_input("cimento_Resistencia_real_7d (MPa):", min_value=0.0, step=1.0, key="resistencia_7d"),
            st.number_input("cimento_Resistencia_real_28d (MPa):", min_value=0.0, step=1.0, key="resistencia_28d"),
            st.number_input("Fc_7d (MPa):", min_value=0.0, step=1.0, key="fc_7d"),
            st.number_input("CT_Silica (kg/m³):", min_value=0.0, step=1.0, key="ct_silica"),
            st.number_input("CT_Plastificante (kg/m³):", min_value=0.0, step=1.0, key="ct_plastificante"),
            st.number_input("CT_Polifuncional (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Superplastificante (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Brita_0 (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Brita_1 (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Areia_natural (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Areia_artificial (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_AC (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Aditivo (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Teor_de_Argamassa (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("CT_Teor_de_Agua (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("Volume (m³):", min_value=0.0, step=1.0),
            st.number_input("Mesp_Brita_0 (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("Mesp_Brita_1 (kg/m³):", min_value=0.0, step=1.0),
            st.number_input("Tempo_de_transporte (s):", min_value=0.0, step=1.0),
            st.number_input("Slump (cm):", min_value=0.0, step=1.0),
        ]

    # Botão para calcular
    if st.button("Calcular Resistência"):
        if all(isinstance(v, (int, float)) and v > 0 for v in st.session_state.entradas):  # Verifica se são números positivos
            try:
                model_mediana = joblib.load(model_path)
                model_inf = joblib.load(model_path.replace("0.500", "0.025")) #Carrega o modelo de quantil inferior
                model_sup = joblib.load(model_path.replace("0.500", "0.975")) #Carrega o modelo de quantil superior

                entrada = np.array([st.session_state.entradas])
                mediana = model_mediana.predict(entrada)[0]
                inferior = model_inf.predict(entrada)[0]
                superior = model_sup.predict(entrada)[0]

                st.success(f"A resistência prevista do concreto aos 28 dias é: **{mediana:.2f} MPa**")
                st.write(f"Intervalo de confiança de 95%: [{inferior:.2f} MPa, {superior:.2f} MPa]")

            except FileNotFoundError:
                st.error("Modelo não encontrado. Verifique o caminho do modelo.")
            except Exception as e:
                st.error(f"Erro ao carregar o modelo ou realizar a predição: {e}")
        else:
            st.error("Por favor, insira valores numéricos positivos válidos para todas as variáveis.")


elif st.session_state.tipo_entrada == "Upload Excel":
    # Criar duas colunas: uma para o texto e outra para o botão
    col1, col2 = st.columns([3, 1])  # Ajuste as proporções conforme necessário

    # Coluna da esquerda com o texto
    with col1:
        st.write(
            "Para fazer o upload do Excel, utilize o modelo ao lado&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"
            "<span style='font-size: 30px; color: #4CAF50;'>&#8594;</span>",
            "<span style='font-size: 30px; color: #4CAF50;'>&#8594;</span>",
            "<span style='font-size: 30px; color: #4CAF50;'>&#8594;</span>",
            "<span style='font-size: 30px; color: #4CAF50;'>&#8594;</span>",
            "<span style='font-size: 30px; color: #4CAF50;'>&#8594;</span>",
            "<span style='font-size: 30px; color: #4CAF50;'>&#8594;</span>",
            "<span style='font-size: 30px; color: #4CAF50;'>&#8594;</span>",
            unsafe_allow_html=True
        )

    # Coluna da direita com o botão de download
    with col2:
        # Adicionar espaço acima do botão
        st.markdown("<div style='margin-top: 0px;'></div>", unsafe_allow_html=True)
        with open("modelo.xlsx", "rb") as f:
            st.download_button(
                label="Baixar modelo",
                data=f,
                file_name="modelo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # Opções de configuração para o Excel
    opcao_excel = st.radio(
        "Escolha a configuração de entrada:",
        [
            "CT_Cimento e CT_Agua",
            "CT_Cimento, CT_Agua, e resistências reais (3d, 7d, 28d)",
            "CT_Cimento, CT_Agua, resistências reais, e Fc_7d",
            "CT_Cimento, CT_Agua, resistências reais, Fc_7d, e aditivos",
            "Todas as variáveis"
        ]
    )

    # Carregar o arquivo Excel
    uploaded_file = st.file_uploader("Escolha um arquivo Excel", type=["xlsx", "xls"])

    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)
    
        try:
            # Seleção do modelo e das variáveis de entrada com base na configuração escolhida
            if opcao_excel == "CT_Cimento e CT_Agua":
                model = joblib.load("modelo1.pkl")
                selected_columns = ["CT_Cimento", "CT_Agua"]
    
            elif opcao_excel == "CT_Cimento, CT_Agua, e resistências reais (3d, 7d, 28d)":
                model = joblib.load("modelo2.pkl")
                selected_columns = ["CT_Cimento", "CT_Agua", "cimento_Resistencia_real_3d", 
                                    "cimento_Resistencia_real_7d", "cimento_Resistencia_real_28d"]
    
            elif opcao_excel == "CT_Cimento, CT_Agua, resistências reais, e Fc_7d":
                model = joblib.load("modelo3.pkl")
                selected_columns = ["CT_Cimento", "CT_Agua", "cimento_Resistencia_real_3d", 
                                    "cimento_Resistencia_real_7d", "cimento_Resistencia_real_28d", "Fc_7d"]
    
            elif opcao_excel == "CT_Cimento, CT_Agua, resistências reais, Fc_7d, e aditivos":
                model = joblib.load("modelo4.pkl")
                selected_columns = ["CT_Cimento", "CT_Agua", "cimento_Resistencia_real_3d", 
                                    "cimento_Resistencia_real_7d", "cimento_Resistencia_real_28d", 
                                    "Fc_7d", "CT_Silica", "CT_Plastificante", "CT_Polifuncional", 
                                    "CT_Superplastificante", "CT_Brita_0", "CT_Brita_1", "CT_Areia_natural", 
                                    "CT_Areia_artificial", "CT_AC", "CT_Aditivo", "CT_Teor_de_Argamassa", 
                                    "CT_Teor_de_Agua"]
    
            elif opcao_excel == "Todas as variáveis":
                model = joblib.load("modelo5.pkl")
                selected_columns = ["CT_Cimento", "CT_Agua", "cimento_Resistencia_real_3d", 
                                    "cimento_Resistencia_real_7d", "cimento_Resistencia_real_28d", 
                                    "Fc_7d", "CT_Silica", "CT_Plastificante", "CT_Polifuncional", 
                                    "CT_Superplastificante", "CT_Brita_0", "CT_Brita_1", "CT_Areia_natural", 
                                    "CT_Areia_artificial", "CT_AC", "CT_Aditivo", "CT_Teor_de_Argamassa", 
                                    "CT_Teor_de_Agua", "Volume", "Mesp_Brita_0", "Mesp_Brita_1", 
                                    "Tempo_de_transporte", "Slump"]

            # Carregar os modelos (mediana, inferior e superior)
            model_mediana = joblib.load(model_path)
            model_inf = joblib.load(model_path.replace("0.500", "0.025"))
            model_sup = joblib.load(model_path.replace("0.500", "0.975"))

            # Selecionar as colunas relevantes e preparar os dados para previsão
            input_data = df[selected_columns]
            X = input_data.values

            # Fazer as previsões para mediana e intervalo
            previsoes_mediana = model_mediana.predict(X)
            previsoes_inferior = model_inf.predict(X)
            previsoes_superior = model_sup.predict(X)

            # Adicionar as previsões ao DataFrame
            input_data["Previsão (Mediana)"] = previsoes_mediana
            input_data["Limite Inferior (95%)"] = previsoes_inferior
            input_data["Limite Superior (95%)"] = previsoes_superior

            # Destaque na tela
            def highlight_column(val):
                # Aplica estilo apenas para a coluna "Previsões"
                return 'background-color: #FFFF00; font-weight: bold;' if val == "Previsões" else None
    
            st.write("Variáveis de entrada acompanhadas das respectivas previsões:")
            st.dataframe(input_data.style.applymap(highlight_column, subset=["Previsões"]))
    
            # Salvar o DataFrame atualizado em um novo arquivo Excel
            output_file = "previsoes_" + uploaded_file.name
            input_data.to_excel(output_file, index=False)
    
            # Aplicar estilo no Excel
            wb = load_workbook(output_file)
            ws = wb.active
    
            # Aplicar destaque à coluna "Previsões"
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)
    
            previsoes_col_idx = len(selected_columns) + 1  # Índice da coluna "Previsões"
            for row in ws.iter_rows(min_row=2, min_col=previsoes_col_idx, max_col=previsoes_col_idx):
                for cell in row:
                    cell.fill = yellow_fill
                    cell.font = bold_font
    
            wb.save(output_file)
    
            # Oferecer o arquivo para download
            with open(output_file, "rb") as f:
                st.download_button(
                    label=("Baixar Excel com as previsões"),
                    data=f,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
        except FileNotFoundError:
            st.error("Modelo não encontrado. Verifique o caminho do modelo.")
        except Exception as e:
            st.error(f"Erro ao realizar a previsão: {e}")
